from datetime import datetime
import re
import json
import openpyxl

fp = open('./归属地.json','r',encoding='utf8')
area_dict = json.load(fp)



def checkArea(areaId):
    '''校验地区。输入6位地区编码字符串，返回：{code, id, {province, city, district, note}}'''
    if len(areaId) != 6:
        return {'code': 'Error', 'id': areaId, 'area': '区域码长度需要为6位'}
    if areaId.isdigit() == False:
        return {'code': 'Error', 'id': areaId, 'area': '非法区域码'}
    if areaId in area_dict:
        return {'code': 'OK', 'id': areaId, 'area': {'province': area_dict[areaId]['p'], 'city': area_dict[areaId]['a'], 'district': area_dict[areaId]['d'], 'note': area_dict[areaId]['i']}}
    else:
        return {'code': 'Error', 'id': areaId, 'area': '未知区域码'}
    


def checkBirthdate(ymd):
    '''校验出生日期。输入8位出生日期字符串，返回：{code, id, age, {year, month, day}}'''
    if len(ymd) != 8:
        return {'code': 'Error', 'id': ymd, 'age': '出生日期长度需要为8位', 'date': ''}
    if ymd.isdigit() == False:
        return {'code': 'Error', 'id': ymd, 'age': '非法出生日期', 'date': ''}
    year = int(ymd[0:4])
    month = int(ymd[4:6])
    day = int(ymd[6:8])
    currentYear = datetime.now().year
    age = currentYear - year
    if age < 0:
        return {'code': 'Error', 'id': ymd, 'age': '未来出生日期', 'date': ''}
    # 年份：[1-9][0-9]{3}
    # 闰年月日:((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|[1-2][0-9]))
    # 平年月日:((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|1[0-9]|2[0-8]))
    # 使用正则表达式
    if (year % 4 == 0 and year % 100 != 0) or year % 400 == 0:
        ereg = re.compile('([1-9][0-9]{3})((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|[1-2][0-9]))')
    else: 
        ereg = re.compile('([1-9][0-9]{3})((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|1[0-9]|2[0-8]))')
    if ereg.match(ymd):
        return {'code': 'OK', 'id': ymd, 'age': age, 'date': {'year': year, 'month': month, 'day': day}}
    return {'code': 'Error', 'id': ymd, 'age': '非法出生日期', 'date': ''}


def checkGender(gender):
    '''校验性别。输入1位性别编码字符串，返回：{code, id, gender}'''
    if len(gender) != 1:
        return {'code': 'Error', 'id': gender, 'gender': '性别编码长度需要为1位'}
    if gender.isdigit() == False:
        return {'code': 'Error', 'id':gender, 'gender': '非法性别编码'}
    genderInt = int(gender)
    if genderInt % 2 == 0:
        return {'code': 'OK', 'id':gender, 'gender': '女'}
    else:
        return {'code': 'OK', 'id':gender, 'gender': '男'}
    

'''
1、将身份证号码前17位数分别乘以不同的系数。从第一位到第十七位的系数分别为：7 9 10 5 8 4 2 1 6 3 7 9 10 5 8 4 2
2、将这17位数字和系数相乘的结果相加
3、用加出来和除以11，看余数是多少
4、余数只可能有0 1 2 3 4 5 6 7 8 9 10这11个数字。其分别对应的校验码为1 0 X 9 8 7 6 5 4 3 2
'''

coe = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
lastNum = ['1', '0', 'X', '9', '8', '7', '6', '5', '4', '3', '2']

def checkNum(id):
    '''校验校验码。输入18位字符串，返回：{code, id, area, age, gender, jym}'''
    if len(id) != 18:
        return {'code': 'Error', 'id': id, 'jym': '身份证号码长度需要为18位'}
    # num是id前17位
    num = id[0:17]
    if num.isdigit() == False:
        return {'code': 'Error', 'id': id, 'jym': '非法身份证号码'}
    sum = 0
    for i in range(17):
        sum += int(num[i]) * coe[i]
    jym = lastNum[sum % 11]
    if id[17] != jym:
        return {'code': 'Error', 'id': id, 'jym': '校验码错误'}
    return {'code': 'OK', 'id': id, 'jym': jym}



import checkArea
from checkArea import checkArea
from checkNum import checkNum
from checkBirthdate import checkBirthdate
from checkGender import checkGender

'''
输入为xlsx文件，第一行是表头，第一列从第二行开始为多个身份证号码，将号码读入。
调用checkID函数，检查身份证是否合法有效，并将结果输出到原来的Excel表中，第二列为整体有效性，第三列为校验码正确性，第四列为省份，第5列为城市，第6列为区县，第7列为归属地，第8列为出生日期，第9列为性别
'''
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# 读取xlsx文件
wb = load_workbook('身份证.xlsx')
ws = wb.active
# 读取身份证号码
idList = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        idList.append(cell.value)

# 检查身份证号码
for i in range(len(idList)):
    id = idList[i]
    # checkArea
    areaID = id[0:6]
    areaResult = checkArea(areaID)
    # checkBirthdate
    birthdate = id[6:14]
    birthdateResult = checkBirthdate(birthdate)
    # checkGender
    gender = id[16:17]
    genderResult = checkGender(gender)
    # checkNum
    numResult = checkNum(id)

    # 输出结果
    if areaResult['code'] == 'OK' and birthdateResult['code'] == 'OK' and genderResult['code'] == 'OK' and numResult['code'] == 'OK':
        res = '有效'
    else:
        res = '【异常】'
    if areaResult['code'] == 'Error':
        res_province = '未知'
        res_city = '未知'
        res_district = '未知'
        res_note = '未知'
    else:
        res_province = areaResult['area']['province']
        res_city = areaResult['area']['city']
        res_district = areaResult['area']['district']
        res_note = areaResult['area']['note']
    if birthdateResult['code'] == 'Error':
        res_date = '未知'
    else:
        res_year = birthdateResult['date']['year']
        res_month = birthdateResult['date']['month']
        res_day = birthdateResult['date']['day']
        res_date = str(res_year)+'年'+str(res_month)+'月'+str(res_day)+'日'
    if genderResult['code'] == 'Error':
        res_gender = '未知'
    else:
        res_gender = genderResult['gender']

    ws.cell(row=i+2, column=2, value=res)
    ws.cell(row=i+2, column=3, value=numResult['code'])
    ws.cell(row=i+2, column=4, value=res_province)
    ws.cell(row=i+2, column=5, value=res_city)
    ws.cell(row=i+2, column=6, value=res_district)
    ws.cell(row=i+2, column=7, value=res_note)
    ws.cell(row=i+2, column=8, value=res_date)
    ws.cell(row=i+2, column=9, value=res_gender)
    
    # ws.cell(row=i+2, column=2, value=res)
    # print(id, areaResult['area']['province'])
    # ws.cell(row=i+2, column=3, value=areaResult['area']['province'])
    # ws.cell(row=i+2, column=4, value=areaResult['area']['city'])
    # ws.cell(row=i+2, column=5, value=areaResult['area']['district'])
    # ws.cell(row=i+2, column=6, value=areaResult['area']['note'])
    # ws.cell(row=i+2, column=7, value=str(birthdateResult['date']['year'])+'年'+str(birthdateResult['date']['month'])+'月'+str(birthdateResult['date']['day'])+'日')
    # ws.cell(row=i+2, column=8, value=genderResult['gender'])

# 保存xlsx文件
wb.save('身份证.xlsx')
