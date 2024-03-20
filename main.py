import checkArea
from checkArea import checkArea
from checkNum import checkNum
from checkBirthdate import checkBirthdate
from checkGender import checkGender

'''
输入为xlsx文件，第一行是表头，第一列从第二行开始为多个身份证号码，将号码读入。
调用checkID函数，检查身份证是否合法有效，并将结果输出到原来的Excel表中，第二列为整体有效性，第三列为校验码正确性，第四列为省份，第5列为城市，第6列为区县，第7列为归属地，第8列为出生日期，第9列为性别
'''
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# 读取xlsx文件
wb = load_workbook('身份证.xlsx')
ws = wb.active
# 读取身份证号码
idList = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        idList.append(cell.value)

# 检查身份证号码
for i in range(len(idList)):
    id = idList[i]
    # checkArea
    areaID = id[0:6]
    areaResult = checkArea(areaID)
    # checkBirthdate
    birthdate = id[6:14]
    birthdateResult = checkBirthdate(birthdate)
    # checkGender
    gender = id[16:17]
    genderResult = checkGender(gender)
    # checkNum
    numResult = checkNum(id)

    # 输出结果
    if areaResult['code'] == 'OK' and birthdateResult['code'] == 'OK' and genderResult['code'] == 'OK' and numResult['code'] == 'OK':
        res = '有效'
    else:
        res = '【异常】'
    if areaResult['code'] == 'Error':
        res_province = '未知'
        res_city = '未知'
        res_district = '未知'
        res_note = '未知'
    else:
        res_province = areaResult['area']['province']
        res_city = areaResult['area']['city']
        res_district = areaResult['area']['district']
        res_note = areaResult['area']['note']
    if birthdateResult['code'] == 'Error':
        res_date = '未知'
    else:
        res_year = birthdateResult['date']['year']
        res_month = birthdateResult['date']['month']
        res_day = birthdateResult['date']['day']
        res_date = str(res_year)+'年'+str(res_month)+'月'+str(res_day)+'日'
    if genderResult['code'] == 'Error':
        res_gender = '未知'
    else:
        res_gender = genderResult['gender']

    ws.cell(row=i+2, column=2, value=res)
    ws.cell(row=i+2, column=3, value=numResult['code'])
    ws.cell(row=i+2, column=4, value=res_province)
    ws.cell(row=i+2, column=5, value=res_city)
    ws.cell(row=i+2, column=6, value=res_district)
    ws.cell(row=i+2, column=7, value=res_note)
    ws.cell(row=i+2, column=8, value=res_date)
    ws.cell(row=i+2, column=9, value=res_gender)
    
# 保存xlsx文件
wb.save('身份证.xlsx')
