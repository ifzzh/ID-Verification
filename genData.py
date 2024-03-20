'''
将xzqh_all.xlsx生成字典，生成python直接可用的字典
打开xzqh_all.xlsx文件，读取数据，第一行是表头，第一列为区域编码，第三列为省级，第四列为地市级，第五列为县区级，第六列为数据来源
'''

import openpyxl
import json

# 打开xzqh_all.xlsx文件，读取数据，第一行是表头，第一列为区域编码，第三列为省级，第四列为地市级，第五列为县区级，第六列为数据来源
wb = openpyxl.load_workbook('xzqh_all.xlsx')
ws = wb.active
area_dict = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    area_dict[row[0].value] = {'p': row[2].value, 'a': row[3].value, 'd': row[4].value, 'i': row[5].value}

# 将字典保存到文件
with open('归属地.json', 'w', encoding='utf8') as fp:
    json.dump(area_dict, fp, ensure_ascii=False)